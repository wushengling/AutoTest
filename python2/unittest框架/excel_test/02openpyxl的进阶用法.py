#coding:utf-8
"""
"""
import openpyxl

wb = openpyxl.load_workbook(r"C:\wsl\AutoTest\python2\unittest框架\excel_test\cases.xlsx")
sh = wb["login"]

#1.表单对象的rows属性
#按行获取表单所有格子中的数据,每一行的数据放在一个元组中
datas = list(sh.rows)
#通过下标获取指定行的数据
# li = []
# for i in datas[0]:
#     # print(i.value)
#     li.append(i.value)
# print(li)

#列表推导式(扩展)
li1 = [i.value for i in datas[0]]
print(li1)
li2 = [i.value for i in datas[1]]
print(li2)
datas = zip(li1,li2)
print(dict(datas))

#2.读取指定列的数据
datas = list(sh.columns)
cl = [i.value for i in datas[0]]
print(cl)

#3.获取最大行 max_row
print(sh.max_row)

#3.获取最大列 max_column
print(sh.max_column)

