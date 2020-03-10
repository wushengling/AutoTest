#coding:utf-8
"""
openpyxl 返回三大对象
三大对象:
工作薄:workbook
表单:sheet
格子:cell

"""
#通过openpyxl这个模块来操作
import openpyxl

#1.数据的读取
#第一步:读取excel文件,返回一个workbook对象(工作簿)
# wb = openpyxl.load_workbook(r"C:\wsl\AutoTest\python2\unittest框架\excel_test\cases.xlsx")
# print(wb)

#第二步:选取工作薄中的表单
# sh = wb["login"]
# print(sh)

#第三步:读取表单中某个格子的内容
# data = sh.cell(row=1,column=1).value
# data1 = sh.cell(row=2,column=1).value
# print(data1)

#2.数据的写入
#第一步:读取excel文件,返回一个workbook对象(工作簿)
wb = openpyxl.load_workbook(r"C:\wsl\AutoTest\python2\unittest框架\excel_test\cases.xlsx")

print(wb)

#第二步:选取工作薄中的表单
sh = wb["Sheet2"]
print(sh)

#第三步:写入内容
sh.cell(row=1,column=2,value="写入的内容")
sh.cell(row=1,column=1).value="写入的内容"

#第四步:保存工作薄为文件
wb.save(r"C:\wsl\AutoTest\python2\unittest框架\demo2\cases.xlsx")
